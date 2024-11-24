import pandas as pd
from openpyxl import load_workbook


file_path = "test.xlsx"

wb=load_workbook(file_path)
sheet_name="Sheet1"
ws=wb[sheet_name]

ws["A1"]="updated 555"
wb.save(file_path)

print ("data updated successfully")